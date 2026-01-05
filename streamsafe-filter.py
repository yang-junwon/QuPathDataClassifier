#!/usr/bin/env python3
"""
streamsafe_cd4_foxp3.py

Streaming-safe processing for large Excel workbooks:
- Identifies CD4+ FOXP3+ phenotype rows (case-insensitive substring match)
- Collects phenotype labels in PASS 1 (streaming)
- Filters ONLY those phenotypes in PASS 2
- Detects a valid distance column
- Splits CD4+FOXP3+ rows into:
    • within ±100 microns
    • outside that range
- Adds subtype sheets (based on Sample Name)
- Fully stream-safe: openpyxl read_only + write_only
"""

from openpyxl import load_workbook, Workbook
import os
import sys

# -------------------- USER CONFIG --------------------
INPUT_FILE  = "consolidated_mIF data analysis_Nov2025_JY.xlsx"
OUTPUT_FILE = "CD4_FOXP3_filtered_output_streamsafe_v2.xlsx"

# Exact Sample Name → Subtype mapping
sample_to_subtype = {
    "8F_morph": "morpheaform",
    "10A_meta": "meta",
    "2D_inf": "infiltrative",
    "5F_micro": "micronodular",
    "8D_mix": "mixed (NI)",
    "9A_nod": "nodular",
    "4B_super": "superficial"
}

# Column detection logic
def is_distance_col_name(colname: str) -> bool:
    if not colname:
        return False
    cl = colname.lower()
    return ("d  istance" in cl) and (
        "micron" in cl or "edge" in cl or "process" in cl or "tissue" in cl
    )

# Excel sheet name limit
MAX_SHEETNAME = 31

def unique_name(base, used):
    base = base[:MAX_SHEETNAME]
    if base not in used:
        used.add(base)
        return base
    for i in range(2, 999):
        suffix = f"_{i}"
        name = base[:MAX_SHEETNAME - len(suffix)] + suffix
        if name not in used:
            used.add(name)
            return name
    raise RuntimeError("Unable to create unique sheet name.")


# -------------------- PASS 1: DETECT CD4+ FOXP3+ LABELS --------------------
print("PASS 1 — scanning workbook for CD4+ FOXP3+ phenotype labels (streaming)...")

# Edit such that it detects cells only CD4/FOXp3 not CD4 solely.
# if it says FOXP3 then it's also CD4/FOXP3
#There are no FOXP3 cells alone in dataset.

if not os.path.exists(INPUT_FILE):
    print(f"Input file not found: {INPUT_FILE}")
    sys.exit(1)

wb_ro = load_workbook(INPUT_FILE, read_only=True, data_only=True)

candidate_labels = set()

for sheetname in wb_ro.sheetnames:
    ws = wb_ro[sheetname]

    # Read header
    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        continue

    header = [(str(h).strip() if h is not None else "") for h in header_row]
    lc_to_idx = {h.lower(): i for i, h in enumerate(header)}

    if "phenotype" not in lc_to_idx:
        continue

    ph_idx = lc_to_idx["phenotype"]

    # Stream through rows
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[ph_idx] if ph_idx < len(row) else None
        if val is None:
            continue
        s = str(val).strip()
        if not s:
            continue

        sl = s.lower()
        if ("cd4" in sl) and ("foxp3" in sl):
            candidate_labels.add(s)

wb_ro.close()

print(f"Detected {len(candidate_labels)} CD4+ FOXP3+ labels:")
for c in sorted(candidate_labels):
    print("  -", c)

if not candidate_labels:
    print("⚠ No phenotype labels found by substring. Script will use fallback substring matching.")


# -------------------- PASS 2: FILTER + WRITE OUTPUT --------------------
print("\nPASS 2 — streaming filter and writing output...")

wb_ro = load_workbook(INPUT_FILE, read_only=True, data_only=True)
wb_w  = Workbook(write_only=True)
used_names = set()

total_rows_written = 0
total_sheets = 0

for sheetname in wb_ro.sheetnames:
    print(f"\nProcessing sheet: {sheetname}")

    ws = wb_ro[sheetname]

    try:
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        print(" Empty sheet — skipped.")
        continue

    header = [(str(h).strip() if h is not None else "") for h in header_row]
    lc_to_idx = {h.lower(): i for i, h in enumerate(header)}

    # Must have phenotype column
    if "phenotype" not in lc_to_idx:
        print(" No Phenotype column — skipping.")
        continue
    ph_idx = lc_to_idx["phenotype"]

    # Find distance column
    dist_idx = None
    dist_name = None

    # Preferred detection
    for idx, colname in enumerate(header):
        if is_distance_col_name(colname):
            dist_idx = idx
            dist_name = colname
            break

    # Fallback: any column containing "distance"
    if dist_idx is None:
        for idx, colname in enumerate(header):
            if colname and "distance" in colname.lower():
                dist_idx = idx
                dist_name = colname
                break

    if dist_idx is None:
        print("  No distance column found — skipping sheet.")
        continue

    print(f"  Using distance column: “{dist_name}”")

    # Sample Name index (optional)
    sample_idx = lc_to_idx.get("sample name", None)

    # Create output sheets
    within_name  = unique_name(f"{sheetname}_CD4_FOXP3_within100", used_names)
    outside_name = unique_name(f"{sheetname}_CD4_FOXP3_outside100", used_names)

    ws_within  = wb_w.create_sheet(within_name)
    ws_outside = wb_w.create_sheet(outside_name)
    total_sheets += 2

    # Write headers
    ws_within.append(header)
    ws_outside.append(header)

    # Subtype sheets (created on demand)
    subtype_sheets = {}

    # -------- STREAM ROWS --------
    for row in ws.iter_rows(min_row=2, values_only=True):

        # Skip empty row
        if row is None or all(v is None for v in row):
            continue

        phen_val = row[ph_idx] if ph_idx < len(row) else None
        if phen_val is None:
            continue

        phen_str = str(phen_val).strip()
        if not phen_str:
            continue

        sl = phen_str.lower()

        # Determine CD4+FOXP3+ (candidate exact match or fallback substring)
        if candidate_labels:
            is_cd4foxp3 = phen_str in candidate_labels
        else:
            is_cd4foxp3 = ("cd4" in sl and "foxp3" in sl)

        if not is_cd4foxp3:
            continue  # skip non-target phenotypes

        # Parse distance
        try:
            d = row[dist_idx]
            if isinstance(d, str):
                d = float(d.strip())
            elif isinstance(d, (int, float)):
                d = float(d)
            else:
                d = float(d)
        except Exception:
            continue

        # Append filtered rows
        if -100 <= d <= 100:
            ws_within.append(list(row))
        else:
            ws_outside.append(list(row))

        total_rows_written += 1

        # Subtype mapping
        if sample_idx is not None:
            sample_val = row[sample_idx]
            if sample_val is not None:
                sample_str = str(sample_val).strip()
                if sample_str in sample_to_subtype:
                    subtype = sample_to_subtype[sample_str]
                    if subtype not in subtype_sheets:
                        sname = unique_name(f"{sheetname}_{subtype}", used_names)
                        s_ws = wb_w.create_sheet(sname)
                        s_ws.append(header)
                        subtype_sheets[subtype] = s_ws
                        total_sheets += 1
                    subtype_sheets[subtype].append(list(row))

print(f"\nSaving output workbook: {OUTPUT_FILE}")
wb_w.save(OUTPUT_FILE)

print(f"Sheets written: {total_sheets}, Rows written: {total_rows_written}")
